"""Pipeline: Wind Measurement Stations Dataset.

Downloads, normalizes, resamples and consolidates wind measurement data
from 43 ground measurement stations in northern Chile into a single
monthly-aggregated dataset.
"""

import re
import time
import urllib.request
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl


ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
EXCEL_PATH = DATA_DIR / "Estaciones de Medición Eólica.xlsx"
CACHE_DIR = DATA_DIR / "wind_stations"
CATALOG_PATH = DATA_DIR / "station_catalog_wind.csv"
DATASET_PATH = DATA_DIR / "dataset_eolico_mensual.csv"
DATASET_MD_PATH = DATA_DIR / "dataset_eolico_mensual.md"

TARGET_WIND_HEIGHT = 20.0


def parse_station_sheet(ws) -> dict:
    """Extract metadata from a single station sheet using positional parsing."""
    code = ""
    name = ""
    region = ""
    dates_str = ""
    lat = None
    lon = None
    elev = None
    link = ""

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2, values_only=True):
        key = str(row[0]).strip() if row[0] is not None else ""

        if "código estación:" in key.lower():
            code = key.split(":", 1)[-1].strip()
        elif "estación:" in key.lower() and "código" not in key.lower():
            name = key.split(":", 1)[-1].strip()
        elif "comuna" in key.lower():
            parts = key.split(",")
            region = parts[-1].strip() if len(parts) >= 2 else key.strip()
        elif "fechas de medición" in key.lower():
            dates_str = key
        elif "latitud:" in key.lower() and row[1] is not None:
            lat = _parse_numeric(row[1])
        elif "longitud:" in key.lower() and row[1] is not None:
            lon = _parse_numeric(row[1])
        elif "elevación:" in key.lower() and row[1] is not None:
            elev_match = re.search(r"(\d+(?:\.\d+)?)", str(row[1]))
            if elev_match:
                elev = float(elev_match.group(1))
        elif "http" in key.lower():
            link = key.strip()

    fecha_inicio, fecha_fin = "", ""
    match = re.search(r"desde\s+(\d{1,2}-\d{1,2}-\d{4})\s+hasta\s+(\d{1,2}-\d{1,2}-\d{4})", dates_str)
    if match:
        fecha_inicio = match.group(1)
        fecha_fin = match.group(2)

    return {
        "codigo": code,
        "nombre": name,
        "region": region,
        "latitud": lat,
        "longitud": lon,
        "elevacion_m": elev,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "link_datos": link,
    }


def _parse_numeric(val):
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def build_catalog() -> pd.DataFrame:
    """Parse the Excel file and return a DataFrame catalog of all stations."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        info = parse_station_sheet(ws)
        records.append(info)
    wb.close()

    df = pd.DataFrame(records)
    df.to_csv(CATALOG_PATH, index=False)
    print(f"[CATALOG] {len(df)} estaciones -> {CATALOG_PATH}")
    return df


def download_csv(url: str, dest: Path, retries: int = 3, delay: float = 2.0) -> bool:
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Yakhchal-DataCenter/0.1"})
            with urllib.request.urlopen(req, timeout=60) as resp:
                data = resp.read()
            dest.write_bytes(data)
            return True
        except Exception as e:
            print(f"  [WARN] Attempt {attempt + 1}/{retries} for {url}: {e}")
            if attempt < retries - 1:
                time.sleep(delay)
    return False


def ensure_all_csvs(catalog: pd.DataFrame) -> dict:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    paths = {}
    for _, row in catalog.iterrows():
        code = row["codigo"]
        url = row["link_datos"]
        dest = CACHE_DIR / f"{code}.csv"
        if dest.exists():
            print(f"[CACHE] {code}.csv ya existe, omitiendo descarga")
        else:
            print(f"[DOWNLOAD] {code}.csv <- {url}")
            success = download_csv(url, dest)
            if not success:
                print(f"  [ERROR] Falló descarga de {code}")
                continue
        paths[code] = dest
    return paths


def _extract_height(col_name: str) -> float | None:
    """Extract anemometer height from a wind column name. Returns float or None."""
    m = re.search(r"en\s+(\d+(?:\.\d+)?)\s*metros", col_name, re.IGNORECASE)
    if m:
        return float(m.group(1))
    return None


def _pick_best_wind_speed(cols: list[str]) -> str | None:
    """From all wind speed mean columns, pick the one with height closest to TARGET_WIND_HEIGHT."""
    candidates = []
    wind_speed_pat = re.compile(r"velocidad.*viento.*\[mean.*m/s", re.IGNORECASE)
    for c in cols:
        if wind_speed_pat.search(c):
            h = _extract_height(c)
            if h is not None:
                candidates.append((abs(h - TARGET_WIND_HEIGHT), h, c))
    if not candidates:
        return None
    candidates.sort()
    return candidates[0][2]


def _find_first_col(cols: list[str], pattern: re.Pattern) -> str | None:
    for c in cols:
        if pattern.search(c):
            return c
    return None


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Detect columns by pattern and rename to canonical names."""
    cols = df.columns.tolist()

    time_col = None
    for c in cols:
        if "fecha" in c.lower() or "date" in c.lower() or "time" in c.lower():
            time_col = c
            break

    best_wind = _pick_best_wind_speed(cols)

    wind_dir_pat = re.compile(r"dirección.*viento.*\[mean.*grados", re.IGNORECASE)
    temp_pat = re.compile(r"temperatura.*\[mean.*[cC]", re.IGNORECASE)
    hum_pat = re.compile(r"humedad", re.IGNORECASE)
    press_pat = re.compile(r"presi[óo]n\s*atmosf[ée]rica", re.IGNORECASE)
    ghi_pat = re.compile(r"global\s+horizontal", re.IGNORECASE)

    wind_dir = _find_first_col(cols, wind_dir_pat)
    temp_col = _find_first_col(cols, temp_pat)
    hum_col = _find_first_col(cols, hum_pat)
    press_col = _find_first_col(cols, press_pat)
    ghi_col = _find_first_col(cols, ghi_pat)

    mapping = {}
    rename = {}
    if best_wind:
        mapping[best_wind] = "wind_speed"
        rename[best_wind] = "wind_speed"
    if wind_dir:
        mapping[wind_dir] = "wind_direction"
        rename[wind_dir] = "wind_direction"
    if temp_col:
        mapping[temp_col] = "temperatura"
        rename[temp_col] = "temperatura"
    if hum_col:
        mapping[hum_col] = "humedad"
        rename[hum_col] = "humedad"
    if press_col:
        mapping[press_col] = "presion"
        rename[press_col] = "presion"
    if ghi_col:
        mapping[ghi_col] = "ghi"
        rename[ghi_col] = "ghi"

    if not mapping:
        return pd.DataFrame()

    subset_cols = [time_col] + list(mapping.keys()) if time_col else list(mapping.keys())
    df_sub = df[subset_cols].copy()
    df_sub = df_sub.rename(columns=rename)
    if time_col:
        df_sub = df_sub.rename(columns={time_col: "timestamp"})

    for var in mapping.values():
        if var in df_sub.columns:
            df_sub[var] = pd.to_numeric(df_sub[var], errors="coerce")

    return df_sub


RENAME_MAP = {
    "wind_speed": "wind_speed_mean_ms",
    "wind_direction": "wind_direction_mean_deg",
    "temperatura": "temperatura_mean_c",
    "humedad": "humedad_mean_pct",
    "presion": "presion_mean_hpa",
    "ghi": "ghi_mean_wm2",
}


def load_and_normalize(filepath: Path) -> pd.DataFrame | None:
    try:
        raw = pd.read_csv(filepath, encoding="utf-8", low_memory=False)
    except UnicodeDecodeError:
        raw = pd.read_csv(filepath, encoding="latin-1", low_memory=False)

    df = normalize_columns(raw)
    if df.empty:
        print(f"  [SKIP] {filepath.name}: no se detectaron columnas de interés")
        return None

    if "timestamp" in df.columns:
        df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce")
        df = df.dropna(subset=["timestamp"])
        df = df.set_index("timestamp").sort_index()
    else:
        print(f"  [WARN] {filepath.name}: no se encontró columna de timestamp")
        return None

    numeric_cols = [c for c in df.columns if c in RENAME_MAP]
    if not numeric_cols:
        return None

    df = df[numeric_cols]
    return df


def resample_monthly(df: pd.DataFrame) -> pd.DataFrame:
    monthly = df.resample("ME").mean()
    monthly = monthly.dropna(how="all")
    monthly = monthly.rename(columns=RENAME_MAP)
    return monthly


SENTINEL_RULES = {
    "temperatura_mean_c": (-30.0, "lt"),
    "ghi_mean_wm2": (1500.0, "gt"),
}


def filter_sentinels(df: pd.DataFrame) -> pd.DataFrame:
    """Replace physically impossible sentinel values with NaN."""
    for col, (threshold, direction) in SENTINEL_RULES.items():
        if col in df.columns:
            if direction == "lt":
                mask = df[col] < threshold
            else:
                mask = df[col] > threshold
            n_filtered = mask.sum()
            if n_filtered > 0:
                df.loc[mask, col] = np.nan
                op = "<" if direction == "lt" else ">"
                print(f"  [FILTER] {col}: {n_filtered} valores centinela ({op} {threshold}) → NaN")
    return df


def consolidate(catalog: pd.DataFrame, cached_paths: dict) -> pd.DataFrame:
    all_records = []
    for _, row in catalog.iterrows():
        code = row["codigo"]
        if code not in cached_paths:
            continue

        filepath = cached_paths[code]
        print(f"[PROCESS] {code} ({row['nombre']})")
        df = load_and_normalize(filepath)
        if df is None or df.empty:
            continue

        monthly = resample_monthly(df)
        if monthly.empty:
            continue

        monthly = filter_sentinels(monthly)

        monthly["estacion"] = code
        monthly["nombre"] = row["nombre"]
        monthly["region"] = row["region"]
        monthly["longitud"] = row["longitud"]
        monthly["latitud"] = row["latitud"]
        monthly["elevacion_m"] = row["elevacion_m"]
        monthly["fuente_datos"] = row["link_datos"]

        monthly = monthly.reset_index().rename(columns={"timestamp": "fecha"})
        all_records.append(monthly)
        print(f"  {len(monthly)} meses procesados ({monthly['fecha'].min().strftime('%Y-%m')} → {monthly['fecha'].max().strftime('%Y-%m')})")

    if not all_records:
        raise RuntimeError("No se procesó ninguna estación")

    final = pd.concat(all_records, ignore_index=True)

    column_order = [
        "estacion", "nombre", "region", "longitud", "latitud",
        "elevacion_m", "fecha", "wind_speed_mean_ms", "wind_direction_mean_deg",
        "temperatura_mean_c", "humedad_mean_pct", "presion_mean_hpa",
        "ghi_mean_wm2", "fuente_datos",
    ]

    for col in column_order:
        if col not in final.columns:
            final[col] = np.nan

    final = final[column_order]
    final = final.sort_values(["estacion", "fecha"]).reset_index(drop=True)

    return final


def generate_metadata_report(dataset: pd.DataFrame) -> str:
    lines = []
    lines.append("# Dataset de Medición Eólica Mensual — Reporte de Metadatos\n")
    lines.append(f"Generado: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}\n")

    lines.append("## Cobertura por Estación\n")
    lines.append("| Estación | Nombre | Región | Lat | Lon | Elev (m) | Inicio | Fin | Meses | Viento | Dir | Temp | HR | Pres | GHI |")
    lines.append("|---|---|---|---|---|---|---|---|---|---|---|---|---|---|---|")

    for estacion, grp in dataset.groupby("estacion"):
        row_data = dataset[dataset["estacion"] == estacion].iloc[0]

        def check(col):
            return "✓" if grp[col].notna().any() else "✗"

        lines.append(
            f"| {estacion} | {row_data['nombre']} | {row_data['region']} "
            f"| {row_data['latitud']:.2f} | {row_data['longitud']:.2f} "
            f"| {row_data['elevacion_m']:.0f} "
            f"| {grp['fecha'].min().strftime('%Y-%m')} | {grp['fecha'].max().strftime('%Y-%m')} "
            f"| {len(grp)} "
            f"| {check('wind_speed_mean_ms')} | {check('wind_direction_mean_deg')} "
            f"| {check('temperatura_mean_c')} | {check('humedad_mean_pct')} "
            f"| {check('presion_mean_hpa')} | {check('ghi_mean_wm2')} |"
        )

    lines.append("\n## Estadísticas Globales\n")
    lines.append(f"- **Estaciones totales:** {dataset['estacion'].nunique()}")
    lines.append(f"- **Registros mensuales totales:** {len(dataset)}")
    lines.append(f"- **Rango temporal:** {dataset['fecha'].min().strftime('%Y-%m')} → {dataset['fecha'].max().strftime('%Y-%m')}")

    for col, label in [
        ("wind_speed_mean_ms", "Velocidad de viento [m/s]"),
        ("wind_direction_mean_deg", "Dirección del viento [°]"),
        ("temperatura_mean_c", "Temperatura [°C]"),
        ("humedad_mean_pct", "Humedad [%]"),
        ("presion_mean_hpa", "Presión atmosférica [hPa]"),
        ("ghi_mean_wm2", "GHI [W/m²]"),
    ]:
        series = dataset[col].dropna()
        if len(series) > 0:
            lines.append(f"- **{label}:** μ = {series.mean():.1f}, σ = {series.std():.1f}, min = {series.min():.1f}, max = {series.max():.1f} (n = {len(series)})")
        else:
            lines.append(f"- **{label}:** sin datos")

    return "\n".join(lines)


def main():
    print("=== Yakhchal DataCenter — Pipeline de Estaciones Eólicas ===\n")

    print("[1/5] Extrayendo catálogo de estaciones desde Excel...")
    catalog = build_catalog()
    if catalog.empty:
        print("[ERROR] Catálogo vacío. Abortando.")
        return

    print("\n[2/5] Descargando CSVs de datos crudos...")
    cached_paths = ensure_all_csvs(catalog)
    if not cached_paths:
        print("[ERROR] No se pudo descargar ningún CSV. Abortando.")
        return

    print("\n[3/5] Normalizando columnas por detección automática...")
    print("\n[4/5] Remuestreando a medias mensuales...")
    print("\n[5/5] Consolidando dataset final...")
    dataset = consolidate(catalog, cached_paths)

    dataset.to_csv(DATASET_PATH, index=False)
    print(f"\n[DATASET] {len(dataset)} registros -> {DATASET_PATH}")

    report = generate_metadata_report(dataset)
    DATASET_MD_PATH.write_text(report, encoding="utf-8")
    print(f"[REPORT] Metadatos -> {DATASET_MD_PATH}")

    print("\nPipeline completado exitosamente, Capitán.")


if __name__ == "__main__":
    main()
