"""Pipeline: Renewable Energy Curtailment Geospatial Dataset.

Aggregates historical curtailment (energy reduction) reports from the
Coordinador Eléctrico Nacional (CEN) and matches each plant to its
geographic coordinates from the IDE Energía shapefiles, consolidating a
single dataset of accumulated curtailment (MWh) for solar and wind plants
in the SEN of Chile (2022 - Abril 2026).
"""

import datetime
import re
from pathlib import Path

import geopandas as gpd
import numpy as np
import openpyxl
import pandas as pd


ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
DATASET_PATH = DATA_DIR / "curtailment_acumulado.csv"
DATASET_MD_PATH = ROOT / "docs" / "datasets" / "curtailment_acumulado.md"

SHEET_TIPO = {
    "Acumulado-Anual-Solar": "Solar",
    "Acumulado-Anual-Eólico": "Eólica",
}

SHAPEFILE_LAYERS = {
    "Solar": "Solares.shp",
    "Eólica": "Eólicas.shp",
}

PREFIXES = ("PE-", "PFV-", "PMGD-")
NOISE_WORDS = {
    "PARQUE",
    "EOLICO",
    "EOLICA",
    "PFV",
    "PMGD",
    "SOLAR",
    "CENTRAL",
    "FV",
    "PE",
}
ROMAN_TO_ARABIC = {
    "I": "1",
    "II": "2",
    "III": "3",
    "IV": "4",
    "V": "5",
    "VI": "6",
    "VII": "7",
    "VIII": "8",
    "IX": "9",
    "X": "10",
}


def find_month_blocks(ws) -> list[dict]:
    """Locate every monthly block in an Acumulado-Anual sheet.

    Each block is preceded by a marker cell (col B) with a datetime, followed
    two rows below by a header row whose last populated cell reads "Total".
    """
    markers = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, datetime.datetime):
            markers.append((r, v))

    blocks = []
    for i, (marker_row, month) in enumerate(markers):
        header_row = marker_row + 2
        total_row = marker_row + 3

        total_col = None
        for c in range(3, ws.max_column + 1):
            if str(ws.cell(row=header_row, column=c).value).strip() == "Total":
                total_col = c
                break
        if total_col is None:
            continue

        next_marker_row = markers[i + 1][0] if i + 1 < len(markers) else ws.max_row + 1
        blocks.append(
            {
                "month": month,
                "first_row": total_row + 1,
                "last_row": next_marker_row - 1,
                "total_col": total_col,
            }
        )
    return blocks


def extract_block_totals(ws, block: dict) -> list[tuple[str, float]]:
    """Extract (planta, total_mwh) pairs for every plant row in a block."""
    records = []
    for r in range(block["first_row"], block["last_row"] + 1):
        name = ws.cell(row=r, column=2).value
        if name is None:
            continue
        name = str(name).strip()
        if not name or name.lower() == "total":
            continue
        value = ws.cell(row=r, column=block["total_col"]).value
        value = float(value) if isinstance(value, (int, float)) else 0.0
        records.append((name, value))
    return records


def aggregate_curtailment() -> pd.DataFrame:
    """Parse all CEN curtailment workbooks and accumulate totals per plant.

    Monthly blocks across all workbooks are processed in chronological order
    (regardless of filename), so newly commissioned plants appear only once
    their first curtailment record shows up in the historical series.
    """
    workbooks = sorted(DATA_DIR.glob("Reducciones-de-Energia-*.xlsx"))
    if not workbooks:
        raise FileNotFoundError(
            f"No se encontraron reportes de reducciones en {DATA_DIR}"
        )

    open_wbs = {}
    tasks = []  # (month, path, sheet_name, block)
    for path in workbooks:
        wb = openpyxl.load_workbook(path, data_only=True)
        open_wbs[path] = wb
        for sheet_name in SHEET_TIPO:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for block in find_month_blocks(ws):
                tasks.append((block["month"], path, sheet_name, block))
        print(f"  [READ] {path.name}")

    tasks.sort(key=lambda t: t[0])

    totals = {}  # (planta, tipo) -> accumulated MWh, insertion order = first chronological appearance
    for _, path, sheet_name, block in tasks:
        ws = open_wbs[path][sheet_name]
        tipo = SHEET_TIPO[sheet_name]
        for planta, value in extract_block_totals(ws, block):
            key = (planta, tipo)
            totals[key] = totals.get(key, 0.0) + value

    for wb in open_wbs.values():
        wb.close()

    df = pd.DataFrame(
        [{"planta": p, "tipo": t, "curtailment_mwh": v} for (p, t), v in totals.items()]
    )
    print(
        f"[CURTAILMENT] {len(df)} centrales únicas, {len(tasks)} bloques mensuales procesados"
    )
    return df


def find_shapefile_sources() -> list[Path]:
    """List available IDE Energía shapefile exports.

    Manual exports from the web viewer are captured per map viewport, so
    successive `data/output*.zip` copies typically cover different, mostly
    non-overlapping sets of plants rather than growing supersets — all of
    them must be merged to reach full coverage.
    """
    candidates = sorted(DATA_DIR.glob("output*.zip"))
    if not candidates:
        raise FileNotFoundError(
            f"No se encontró ningún export de shapefiles (data/output*.zip) en {DATA_DIR}"
        )
    return candidates


def normalize_name(name: str) -> str:
    """Normalize a plant name/code for cross-source matching.

    Uppercases, strips known CEN prefixes, converts trailing Roman numerals
    to Arabic digits, removes noise terms and non-alphanumeric characters.
    """
    s = name.upper().strip()
    for prefix in PREFIXES:
        if s.startswith(prefix):
            s = s[len(prefix) :]
            break

    tokens = [t for t in re.split(r"[^A-Z0-9]+", s) if t]
    tokens = [ROMAN_TO_ARABIC.get(t, t) for t in tokens]
    tokens = [t for t in tokens if t not in NOISE_WORDS]
    return "".join(tokens)


def load_plant_geometries(zip_paths: list[Path]) -> pd.DataFrame:
    """Load the Solar + Eólica shapefile layers from every export and merge.

    Different exports are deduplicated by (tipo, nombre) so a plant present
    in more than one viewport capture is only counted once.
    """
    records = []
    for zip_path in zip_paths:
        n_before = len(records)
        for tipo, shp_name in SHAPEFILE_LAYERS.items():
            try:
                gdf = gpd.read_file(f"zip://{zip_path}!zipfolder/{shp_name}")
            except Exception as e:
                print(f"  [WARN] {zip_path.name}/{shp_name}: {e}")
                continue
            gdf = gdf.to_crs(epsg=4326)
            for _, row in gdf.iterrows():
                name = str(row["NOMBRE"]).strip()
                records.append(
                    {
                        "tipo": tipo,
                        "nombre_shp": name,
                        "normalizado": normalize_name(name),
                        "longitud": row.geometry.x,
                        "latitud": row.geometry.y,
                    }
                )
        print(f"  [READ] {zip_path.name}: {len(records) - n_before} centrales")

    df = pd.DataFrame(records)
    df = df.drop_duplicates(subset=["tipo", "nombre_shp"], keep="first").reset_index(
        drop=True
    )

    n_solar = (df["tipo"] == "Solar").sum()
    n_eolica = (df["tipo"] == "Eólica").sum()
    print(
        f"[GEOMETRÍAS] {len(df)} centrales únicas tras fusionar exports ({n_solar} solares, {n_eolica} eólicas)"
    )
    return df


def match_plants(curtailment: pd.DataFrame, geometries: pd.DataFrame) -> pd.DataFrame:
    """Match CEN plant codes to shapefile coordinates by normalized name.

    Phase 1: exact match of normalized strings.
    Phase 2: substring containment (either direction) for compound names.
    """
    curtailment = curtailment.copy()
    curtailment["normalizado"] = curtailment["planta"].apply(normalize_name)

    lon, lat, matched_count = [], [], 0
    for _, row in curtailment.iterrows():
        candidates = geometries[geometries["tipo"] == row["tipo"]]
        code = row["normalizado"]
        match = None

        if code:
            exact = candidates[candidates["normalizado"] == code]
            if not exact.empty:
                match = exact.iloc[0]
            else:
                contains = candidates[
                    candidates["normalizado"].apply(
                        lambda n: bool(n) and (code in n or n in code)
                    )
                ]
                if not contains.empty:
                    contains = contains.assign(
                        _diff=(contains["normalizado"].str.len() - len(code)).abs()
                    ).sort_values("_diff")
                    match = contains.iloc[0]

        if match is not None:
            lon.append(match["longitud"])
            lat.append(match["latitud"])
            matched_count += 1
        else:
            lon.append(np.nan)
            lat.append(np.nan)

    curtailment["longitud"] = lon
    curtailment["latitud"] = lat
    print(
        f"[MATCHING] {matched_count}/{len(curtailment)} centrales georreferenciadas ({matched_count / len(curtailment):.0%})"
    )
    return curtailment.drop(columns="normalizado")


def generate_metadata_report(dataset: pd.DataFrame) -> str:
    """Generate a Markdown report with coverage statistics."""
    lines = []
    lines.append("# Dataset de Curtailment Acumulado — Reporte de Metadatos\n")
    lines.append(f"Generado: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}\n")

    total = len(dataset)
    matched = int(dataset["longitud"].notna().sum())
    lines.append("## Cobertura de Georreferenciación\n")
    lines.append(f"- **Centrales totales:** {total}")
    lines.append(f"- **Georreferenciadas:** {matched} ({matched / total:.0%})")
    lines.append(
        f"- **Sin coordenadas:** {total - matched} ({(total - matched) / total:.0%})"
    )

    lines.append("\n## Desglose por Tecnología\n")
    lines.append("| Tipo | Centrales | Georreferenciadas | Curtailment total (MWh) |")
    lines.append("|---|---|---|---|")
    for tipo, grp in dataset.groupby("tipo"):
        geo = int(grp["longitud"].notna().sum())
        lines.append(
            f"| {tipo} | {len(grp)} | {geo} | {grp['curtailment (MWh)'].sum():,.1f} |"
        )

    lines.append("\n## Centrales sin Coordenadas\n")
    unmatched = dataset[dataset["longitud"].isna()][["planta", "tipo"]]
    if unmatched.empty:
        lines.append("Ninguna — cobertura del 100%.")
    else:
        lines.append("| Planta | Tipo |")
        lines.append("|---|---|")
        for _, row in unmatched.iterrows():
            lines.append(f"| {row['planta']} | {row['tipo']} |")

    return "\n".join(lines)


def main():
    print("=== Yakhchal DataCenter — Pipeline de Curtailment Geoespacial ===\n")

    print("[1/4] Parseando reportes de reducciones de energía (CEN)...")
    curtailment = aggregate_curtailment()

    print("\n[2/4] Cargando y fusionando geometrías de centrales (IDE Energía)...")
    zip_paths = find_shapefile_sources()
    geometries = load_plant_geometries(zip_paths)

    print("\n[3/4] Emparejando centrales por nombre normalizado...")
    dataset = match_plants(curtailment, geometries)
    dataset = dataset.rename(columns={"curtailment_mwh": "curtailment (MWh)"})
    dataset = dataset[["planta", "tipo", "longitud", "latitud", "curtailment (MWh)"]]

    print("\n[4/4] Consolidando dataset final...")
    dataset.to_csv(DATASET_PATH, index=False)
    print(f"\n[DATASET] {len(dataset)} registros -> {DATASET_PATH}")

    report = generate_metadata_report(dataset)
    DATASET_MD_PATH.write_text(report, encoding="utf-8")
    print(f"[REPORT] Metadatos -> {DATASET_MD_PATH}")

    print("\nPipeline completado exitosamente, Capitán.")


if __name__ == "__main__":
    main()
