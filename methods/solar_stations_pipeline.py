"""Pipeline: Solar Radiation Measurement Stations Dataset.

Downloads, normalizes, resamples and consolidates solar radiation data
from 11 ground measurement stations in northern Chile into a single
monthly-aggregated dataset.
"""

import re
import time
import urllib.request
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl


ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
EXCEL_PATH = DATA_DIR / "Estaciones de Medición Solar.xlsx"
CACHE_DIR = DATA_DIR / "solar_stations"
CATALOG_PATH = DATA_DIR / "station_catalog.csv"
DATASET_PATH = DATA_DIR / "dataset_solar_mensual.csv"
DATASET_MD_PATH = ROOT / "docs" / "datasets" / "dataset_solar_mensual.md"


def parse_station_sheet(ws) -> dict:
    """Extract metadata from a single station sheet using positional parsing."""
    row_data = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2, values_only=True):
        key = str(row[0]).strip() if row[0] is not None else ""
        val = str(row[1]).strip() if row[1] is not None else ""
        row_data[key.lower()] = val

    code = ""
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2, values_only=True):
        if r[0] and "código estación:" in str(r[0]).lower():
            code = str(r[0]).split(":", 1)[-1].strip()
            break

    name = ""
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2, values_only=True):
        if (
            r[0]
            and "estación:" in str(r[0]).lower()
            and "código" not in str(r[0]).lower()
        ):
            name = str(r[0]).split(":", 1)[-1].strip()
            break

    region = ""
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2, values_only=True):
        if r[0] and "comuna" in str(r[0]).lower():
            raw = str(r[0])
            parts = raw.split(",")
            if len(parts) >= 2:
                region = parts[-1].strip()
            else:
                region = raw.strip()
            break

    dates_str = ""
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2, values_only=True):
        if r[0] and "fechas de medición" in str(r[0]).lower():
            dates_str = str(r[0])
            break

    fecha_inicio, fecha_fin = "", ""
    match = re.search(
        r"desde\s+(\d{1,2}-\d{1,2}-\d{4})\s+hasta\s+(\d{1,2}-\d{1,2}-\d{4})", dates_str
    )
    if match:
        fecha_inicio = match.group(1)
        fecha_fin = match.group(2)

    lat = None
    lon = None
    elev = None
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2, values_only=True):
        if r[0] and "latitud:" in str(r[0]).lower():
            lat = _parse_numeric(r[1]) if r[1] is not None else None
        if r[0] and "longitud:" in str(r[0]).lower():
            lon = _parse_numeric(r[1]) if r[1] is not None else None
        if r[0] and "elevación:" in str(r[0]).lower():
            if r[1] is not None:
                elev_match = re.search(r"(\d+(?:\.\d+)?)", str(r[1]))
                if elev_match:
                    elev = float(elev_match.group(1))

    link = ""
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2, values_only=True):
        if r[0] and "http" in str(r[0]).lower():
            link = str(r[0]).strip()
            break

    return {
        "codigo": code,
        "nombre": name,
        "region": region,
        "latitud": lat,
        "longitud": lon,
        "elevacion_m": elev,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "link_datos": link,
    }


def _parse_numeric(val):
    """Try to parse a numeric value, return None on failure."""
    try:
        return float(val)
    except ValueError, TypeError:
        return None


def build_catalog() -> pd.DataFrame:
    """Parse the Excel file and return a DataFrame catalog of all stations."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        info = parse_station_sheet(ws)
        records.append(info)
    wb.close()

    df = pd.DataFrame(records)
    df.to_csv(CATALOG_PATH, index=False)
    print(f"[CATALOG] {len(df)} estaciones -> {CATALOG_PATH}")
    return df


def download_csv(url: str, dest: Path, retries: int = 3, delay: float = 2.0) -> bool:
    """Download a CSV file to dest, with retries. Returns True on success."""
    for attempt in range(retries):
        try:
            req = urllib.request.Request(
                url, headers={"User-Agent": "Yakhchal-DataCenter/0.1"}
            )
            with urllib.request.urlopen(req, timeout=60) as resp:
                data = resp.read()
            dest.write_bytes(data)
            return True
        except Exception as e:
            print(f"  [WARN] Attempt {attempt + 1}/{retries} for {url}: {e}")
            if attempt < retries - 1:
                time.sleep(delay)
    return False


def ensure_all_csvs(catalog: pd.DataFrame) -> dict:
    """Download all station CSVs to cache dir. Returns {code: filepath}."""
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    paths = {}
    for _, row in catalog.iterrows():
        code = row["codigo"]
        url = row["link_datos"]
        dest = CACHE_DIR / f"{code}.csv"
        if dest.exists():
            print(f"[CACHE] {code}.csv ya existe, omitiendo descarga")
        else:
            print(f"[DOWNLOAD] {code}.csv <- {url}")
            success = download_csv(url, dest)
            if not success:
                print(f"  [ERROR] Falló descarga de {code}")
                continue
        paths[code] = dest
    return paths


COLUMN_PATTERNS = {
    "ghi": re.compile(r"global\s+horizontal", re.IGNORECASE),
    "dni": re.compile(r"directa\s+normal", re.IGNORECASE),
    "temperatura": re.compile(r"temperatura", re.IGNORECASE),
    "humedad": re.compile(r"humedad", re.IGNORECASE),
    "viento": re.compile(r"viento.*mean(?!.*(?:min|max))", re.IGNORECASE),
}

RENAME_MAP = {
    "ghi": "ghi_mean_wm2",
    "dni": "dni_mean_wm2",
    "temperatura": "temperatura_mean_c",
    "humedad": "humedad_mean_pct",
    "viento": "viento_mean_ms",
}


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Detect columns by substring matching and rename to canonical names."""

    def _find_col(cols, pattern):
        for c in cols:
            if pattern.search(c):
                return c
        return None

    cols = df.columns.tolist()
    mapping = {}
    for var, pat in COLUMN_PATTERNS.items():
        matched = _find_col(cols, pat)
        if matched:
            mapping[matched] = var

    if not mapping:
        return pd.DataFrame()

    time_col = None
    for c in cols:
        if "fecha" in c.lower() or "date" in c.lower() or "time" in c.lower():
            time_col = c
            break

    subset_cols = (
        [time_col] + list(mapping.keys()) if time_col else list(mapping.keys())
    )
    df_sub = df[subset_cols].copy()

    df_sub = df_sub.rename(columns=mapping)
    if time_col:
        df_sub = df_sub.rename(columns={time_col: "timestamp"})

    for var in mapping.values():
        if var in df_sub.columns:
            df_sub[var] = pd.to_numeric(df_sub[var], errors="coerce")

    return df_sub


def load_and_normalize(filepath: Path) -> pd.DataFrame | None:
    """Load a station CSV, detect encoding, and normalize columns."""
    try:
        raw = pd.read_csv(filepath, encoding="utf-8", low_memory=False)
    except UnicodeDecodeError:
        raw = pd.read_csv(filepath, encoding="latin-1", low_memory=False)

    df = normalize_columns(raw)
    if df.empty:
        print(f"  [SKIP] {filepath.name}: no se detectaron columnas de interés")
        return None

    if "timestamp" in df.columns:
        df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce")
        df = df.dropna(subset=["timestamp"])
        df = df.set_index("timestamp").sort_index()
    else:
        print(f"  [WARN] {filepath.name}: no se encontró columna de timestamp")
        return None

    numeric_cols = [c for c in df.columns if c in RENAME_MAP]
    if not numeric_cols:
        return None

    df = df[numeric_cols]

    # --- QC FILTERS (Instantaneous values) ---
    if "ghi" in df.columns:
        df.loc[(df["ghi"] < 0) | (df["ghi"] > 1500), "ghi"] = np.nan
    if "dni" in df.columns:
        df.loc[(df["dni"] < 0) | (df["dni"] > 1500), "dni"] = np.nan
    if "viento" in df.columns:
        df.loc[(df["viento"] < 0) | (df["viento"] > 40), "viento"] = np.nan
    if "temperatura" in df.columns:
        df.loc[(df["temperatura"] < -15) | (df["temperatura"] > 45), "temperatura"] = (
            np.nan
        )
    if "humedad" in df.columns:
        df.loc[(df["humedad"] < 0) | (df["humedad"] > 100), "humedad"] = np.nan

    return df


def resample_monthly(df: pd.DataFrame) -> pd.DataFrame:
    """Resample to monthly means. Drops months that are entirely NaN.

    Decisión documentada: Los promedios mensuales de GHI y DNI se calculan sobre
    las 24 horas del día (no solo horas diurnas) para ser consistentes con la
    integración de energía mensual total.
    Se exige un 60% de completitud de datos horarios válidos por mes.
    """
    monthly_mean = df.resample("ME").mean()
    monthly_count = df.resample("ME").count()

    # Requisito de completitud: 60% de horas del mes
    expected_hours = monthly_mean.index.days_in_month * 24
    min_hours = expected_hours * 0.60

    for col in monthly_mean.columns:
        if col in monthly_count.columns:
            monthly_mean.loc[monthly_count[col] < min_hours, col] = np.nan

    monthly = monthly_mean.dropna(how="all")
    monthly = monthly.rename(columns=RENAME_MAP)
    return monthly


def consolidate(catalog: pd.DataFrame, cached_paths: dict) -> pd.DataFrame:
    """Load, normalize, resample, and consolidate all stations."""
    all_records = []
    for _, row in catalog.iterrows():
        code = row["codigo"]
        if code not in cached_paths:
            continue

        filepath = cached_paths[code]
        print(f"[PROCESS] {code} ({row['nombre']})")
        df = load_and_normalize(filepath)
        if df is None or df.empty:
            continue

        monthly = resample_monthly(df)
        if monthly.empty:
            continue

        monthly["estacion"] = code
        monthly["nombre"] = row["nombre"]
        monthly["region"] = row["region"]
        monthly["longitud"] = row["longitud"]
        monthly["latitud"] = row["latitud"]
        monthly["elevacion_m"] = row["elevacion_m"]
        monthly["fuente_datos"] = row["link_datos"]

        monthly = monthly.reset_index().rename(columns={"timestamp": "fecha"})
        all_records.append(monthly)
        print(
            f"  {len(monthly)} meses procesados ({monthly['fecha'].min().strftime('%Y-%m')} → {monthly['fecha'].max().strftime('%Y-%m')})"
        )

    if not all_records:
        raise RuntimeError("No se procesó ninguna estación")

    final = pd.concat(all_records, ignore_index=True)

    column_order = [
        "estacion",
        "nombre",
        "region",
        "longitud",
        "latitud",
        "elevacion_m",
        "fecha",
        "ghi_mean_wm2",
        "dni_mean_wm2",
        "temperatura_mean_c",
        "humedad_mean_pct",
        "viento_mean_ms",
        "fuente_datos",
    ]

    for col in column_order:
        if col not in final.columns:
            final[col] = np.nan

    final = final[column_order]
    final = final.sort_values(["estacion", "fecha"]).reset_index(drop=True)

    return final


def generate_metadata_report(dataset: pd.DataFrame) -> str:
    """Generate a Markdown report with dataset statistics."""
    lines = []
    lines.append("# Dataset de Radiación Solar Mensual — Reporte de Metadatos\n")
    lines.append(f"Generado: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}\n")

    lines.append("## Cobertura por Estación\n")
    lines.append(
        "| Estación | Nombre | Región | Lat | Lon | Elev (m) | Inicio | Fin | Meses | GHI | DNI | Temp | HR | Viento |"
    )
    lines.append("|---|---|---|---|---|---|---|---|---|---|---|---|---|")

    for estacion, grp in dataset.groupby("estacion"):
        row_data = dataset[dataset["estacion"] == estacion].iloc[0]
        has_ghi = grp["ghi_mean_wm2"].notna().any()
        has_dni = grp["dni_mean_wm2"].notna().any()
        has_temp = grp["temperatura_mean_c"].notna().any()
        has_hr = grp["humedad_mean_pct"].notna().any()
        has_wind = grp["viento_mean_ms"].notna().any()

        def check(val):
            return "✓" if val else "✗"

        lines.append(
            f"| {estacion} | {row_data['nombre']} | {row_data['region']} "
            f"| {row_data['latitud']:.2f} | {row_data['longitud']:.2f} "
            f"| {row_data['elevacion_m']:.0f} "
            f"| {grp['fecha'].min().strftime('%Y-%m')} | {grp['fecha'].max().strftime('%Y-%m')} "
            f"| {len(grp)} "
            f"| {check(has_ghi)} | {check(has_dni)} | {check(has_temp)} | {check(has_hr)} | {check(has_wind)} |"
        )

    lines.append("\n## Estadísticas Globales\n")
    lines.append(f"- **Estaciones totales:** {dataset['estacion'].nunique()}")
    lines.append(f"- **Registros mensuales totales:** {len(dataset)}")
    lines.append(
        f"- **Rango temporal:** {dataset['fecha'].min().strftime('%Y-%m')} → {dataset['fecha'].max().strftime('%Y-%m')}"
    )

    for col, label in [
        ("ghi_mean_wm2", "GHI [W/m²]"),
        ("dni_mean_wm2", "DNI [W/m²]"),
        ("temperatura_mean_c", "Temperatura [°C]"),
        ("humedad_mean_pct", "Humedad [%]"),
        ("viento_mean_ms", "Viento [m/s]"),
    ]:
        series = dataset[col].dropna()
        if len(series) > 0:
            lines.append(
                f"- **{label}:** μ = {series.mean():.1f}, σ = {series.std():.1f}, min = {series.min():.1f}, max = {series.max():.1f} (n = {len(series)})"
            )
        else:
            lines.append(f"- **{label}:** sin datos")

    return "\n".join(lines)


def main():
    print("=== Yakhchal DataCenter — Pipeline de Estaciones Solares ===\n")

    print("[1/5] Extrayendo catálogo de estaciones desde Excel...")
    catalog = build_catalog()
    if catalog.empty:
        print("[ERROR] Catálogo vacío. Abortando.")
        return

    print("\n[2/5] Descargando CSVs de datos crudos...")
    cached_paths = ensure_all_csvs(catalog)
    if not cached_paths:
        print("[ERROR] No se pudo descargar ningún CSV. Abortando.")
        return

    print("\n[3/5] Normalizando columnas por detección automática...")
    print("\n[4/5] Remuestreando a medias mensuales...")
    print("\n[5/5] Consolidando dataset final...")
    dataset = consolidate(catalog, cached_paths)

    dataset.to_csv(DATASET_PATH, index=False)
    print(f"\n[DATASET] {len(dataset)} registros -> {DATASET_PATH}")

    report = generate_metadata_report(dataset)
    DATASET_MD_PATH.write_text(report, encoding="utf-8")
    print(f"[REPORT] Metadatos -> {DATASET_MD_PATH}")

    print("\nPipeline completado exitosamente, Capitán.")


if __name__ == "__main__":
    main()
